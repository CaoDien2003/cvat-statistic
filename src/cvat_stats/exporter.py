import csv
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
LABELED_FILL = PatternFill("solid", fgColor="E2EFDA")
UNLABELED_FILL = PatternFill("solid", fgColor="FCE4D6")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        width = min_width
        for cell in col:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def _style_header_row(ws, row_num: int, ncols: int, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_summary_sheet(wb, summary_df: pd.DataFrame):
    ws = wb.create_sheet("Summary")
    is_wide = "value" not in summary_df.columns

    if is_wide:
        # Multi-project: [metric, proj1, proj2, ..., total, percent]
        cols = list(summary_df.columns)
        ws.append(cols)
        _style_header_row(ws, 1, len(cols))
        for i, (_, row) in enumerate(summary_df.iterrows(), start=2):
            ws.append([row.get(c, "") for c in cols])
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            for col_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=i, column=col_idx)
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = LEFT if col_name == "metric" else CENTER
    else:
        has_pct = "percent" in summary_df.columns
        headers = ["Metric", "Value", "Percent (vs total images)"] if has_pct else ["Metric", "Value"]
        ncols = len(headers)
        ws.append(headers)
        _style_header_row(ws, 1, ncols)
        for i, (_, row) in enumerate(summary_df.iterrows(), start=2):
            if has_pct:
                ws.append([row["metric"], row["value"], row.get("percent", "")])
            else:
                ws.append([row["metric"], row["value"]])
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            for col in range(1, ncols + 1):
                cell = ws.cell(row=i, column=col)
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = CENTER if (has_pct and col == 3) else LEFT

    _auto_width(ws)


def _write_label_stats_sheet(wb, label_stats_df: pd.DataFrame):
    ws = wb.create_sheet("Label_Stats")
    is_wide = "images_with_label" not in label_stats_df.columns

    if is_wide:
        # Wide: [class, proj1_imgs, proj1_bbox, proj2_imgs, proj2_bbox, ..., total_imgs, total_bbox, pct_of_labeled]
        proj_names = [c[:-5] for c in label_stats_df.columns if c.endswith("_imgs") and c != "total_imgs"]
        data_cols = ["class"]
        for n in proj_names:
            data_cols += [f"{n}_imgs", f"{n}_bbox"]
        data_cols += ["total_imgs", "total_bbox", "pct_of_labeled"]
    else:
        data_cols = ["class", "images_with_label", "total_bboxes", "pct_of_labeled_imgs"]

    ws.append(data_cols)
    _style_header_row(ws, 1, len(data_cols))

    current_row = 2
    alt = False
    for type_name in ["tag", "polygon", "rectangle"]:
        group = label_stats_df[label_stats_df["type"] == type_name]
        if group.empty:
            continue
        ws.append([type_name.upper()] + [""] * (len(data_cols) - 1))
        _style_header_row(ws, current_row, len(data_cols), fill=SUBHEADER_FILL)
        current_row += 1

        for _, row in group.iterrows():
            ws.append([row.get(c, 0) for c in data_cols])
            fill = ALT_FILL if alt else WHITE_FILL
            alt = not alt
            for col_idx, col_name in enumerate(data_cols, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = LEFT if col_name == "class" else CENTER
            current_row += 1

    # Total row
    total_row = ["TOTAL"]
    for col in data_cols[1:]:
        if "pct" in col:
            total_row.append("")
        elif col in label_stats_df.columns:
            try:
                total_row.append(int(label_stats_df[col].sum()))
            except (TypeError, ValueError):
                total_row.append("")
        else:
            total_row.append("")
    ws.append(total_row)
    _style_header_row(ws, current_row, len(data_cols), fill=SUBHEADER_FILL)

    _auto_width(ws)


def _write_custom_group_sheet(wb, group_df: pd.DataFrame):
    """Render custom user-defined group stats."""
    ws = wb.create_sheet("Group_Stats")
    if group_df.empty:
        ws.append(["No groups defined. Add groups in section 2.5 of the app."])
        return

    cols = list(group_df.columns)
    ws.append(cols)
    _style_header_row(ws, 1, len(cols))

    for i, (_, row) in enumerate(group_df.iterrows(), start=2):
        ws.append([row.get(c, "") for c in cols])
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_name in ("group", "required_labels") else CENTER

    # Total row
    total_row = []
    for col in cols:
        if col in ("group", "required_labels", "percent"):
            total_row.append("TOTAL" if col == "group" else "")
        else:
            try:
                total_row.append(int(group_df[col].sum()))
            except (TypeError, ValueError):
                total_row.append("")
    ws.append(total_row)
    _style_header_row(ws, len(group_df) + 2, len(cols), fill=SUBHEADER_FILL)

    _auto_width(ws)


def _write_group_stats_sheet(wb, group_df: pd.DataFrame):
    # Detect custom group format (has "group" column) → delegate
    if "group" in group_df.columns:
        _write_custom_group_sheet(wb, group_df)
        return

    ws = wb.create_sheet("Group_Stats")
    cols = list(group_df.columns)
    ws.append(cols)
    _style_header_row(ws, 1, len(cols))

    def _subtotal_row(grp, prefix):
        row = [prefix[0], prefix[1]]
        for col in cols[2:]:
            if "pct" in col:
                row.append("")
            elif col in grp.columns:
                try:
                    row.append(int(grp[col].sum()))
                except (TypeError, ValueError):
                    row.append("")
            else:
                row.append("")
        return row

    current_row = 2
    for line_type, grp in group_df.groupby("line_type", sort=False):
        ws.append([line_type.upper()] + [""] * (len(cols) - 1))
        _style_header_row(ws, current_row, len(cols), fill=SUBHEADER_FILL)
        current_row += 1

        alt = False
        for _, row in grp.iterrows():
            ws.append([row.get(c, "") for c in cols])
            fill = ALT_FILL if alt else WHITE_FILL
            alt = not alt
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = LEFT if col_idx <= 2 else CENTER
            current_row += 1

        ws.append(_subtotal_row(grp, ["", "subtotal"]))
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = YELLOW_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_idx <= 2 else CENTER
        current_row += 1

    ws.append(_subtotal_row(group_df, ["TOTAL", ""]))
    _style_header_row(ws, current_row, len(cols), fill=SUBHEADER_FILL)

    _auto_width(ws)


def _write_image_detail_sheet(wb, img_df: pd.DataFrame, labels: list):
    ws = wb.create_sheet("Image_Detail")
    fixed_cols = ["img_name", "task_name", "assignee", "is_labeled"]
    all_cols = fixed_cols + labels
    ws.append(all_cols)
    _style_header_row(ws, 1, len(all_cols))

    for i, (_, row) in enumerate(img_df.iterrows(), start=2):
        values = [row.get(c, "") for c in all_cols]
        ws.append(values)
        base_fill = LABELED_FILL if row.get("is_labeled", 0) else UNLABELED_FILL
        for col_idx in range(1, len(all_cols) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.fill = base_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx > len(fixed_cols) else LEFT

    ws.freeze_panes = "E2"
    _auto_width(ws, max_width=60)


def _write_user_progress_sheet(wb, user_df: pd.DataFrame, labels: list):
    ws = wb.create_sheet("User_Progress")
    fixed_cols = ["assignee", "assigned", "labeled", "unlabeled"]
    all_cols = fixed_cols + labels
    ws.append(all_cols)
    _style_header_row(ws, 1, len(all_cols))

    for i, (_, row) in enumerate(user_df.iterrows(), start=2):
        values = [row.get(c, 0) for c in all_cols]
        ws.append(values)
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for col_idx in range(1, len(all_cols) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx > 1 else LEFT

    total_row_idx = len(user_df) + 2
    ws.append(
        ["TOTAL"]
        + [int(user_df[c].sum()) if c in user_df.columns else "" for c in fixed_cols[1:] + labels]
    )
    _style_header_row(ws, total_row_idx, len(all_cols), fill=SUBHEADER_FILL)

    ws.freeze_panes = "B2"
    _auto_width(ws)


def _write_daily_delta_sheet(wb, delta_df: pd.DataFrame):
    ws = wb.create_sheet("Daily_Delta")
    cols = list(delta_df.columns)
    ws.append(cols)
    _style_header_row(ws, 1, len(cols))

    for i, (_, row) in enumerate(delta_df.iterrows(), start=2):
        ws.append([row.get(c, 0) for c in cols])
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col_name in ("assignee", "prev_date", "note") else CENTER
            if col_name == "delta_labeled":
                val = row.get(col_name, 0)
                if isinstance(val, (int, float)) and val > 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif isinstance(val, (int, float)) and val < 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

    _auto_width(ws)


def export_excel(
    output_path: Path,
    summary_df: pd.DataFrame,
    label_stats_df: pd.DataFrame,
    group_df: pd.DataFrame,
    img_df: pd.DataFrame,
    user_df: pd.DataFrame,
    delta_df: pd.DataFrame,
    labels: list,
):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, summary_df)
    _write_label_stats_sheet(wb, label_stats_df)
    _write_group_stats_sheet(wb, group_df)
    _write_image_detail_sheet(wb, img_df, labels)
    _write_user_progress_sheet(wb, user_df, labels)
    _write_daily_delta_sheet(wb, delta_df)

    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    return output_path


def export_project_csv(output_path: Path, img_df: pd.DataFrame, labels: list):
    cols = ["img_name", "task_name", "assignee", "is_labeled"] + labels
    img_df[cols].to_csv(output_path, index=False, encoding="utf-8-sig")


def append_daily_worker_row(csv_path: Path, run_date: date, project_name: str,
                             user_df: pd.DataFrame):
    cols = ["date", "project", "assignee", "assigned", "labeled", "unlabeled", "pct_labeled"]
    write_header = not csv_path.exists()
    csv_path.parent.mkdir(exist_ok=True)

    with csv_path.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        if write_header:
            writer.writeheader()

        for _, row in user_df.iterrows():
            labeled = int(row.get("labeled", 0))
            assigned = int(row.get("assigned", 0))
            pct = f"{labeled / assigned * 100:.1f}%" if assigned else "0.0%"
            writer.writerow({
                "date": run_date.isoformat(),
                "project": project_name,
                "assignee": row.get("assignee", ""),
                "assigned": assigned,
                "labeled": labeled,
                "unlabeled": int(row.get("unlabeled", 0)),
                "pct_labeled": pct,
            })
